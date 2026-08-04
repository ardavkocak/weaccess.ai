# -*- coding: utf-8 -*-
"""Belge listesini biçimlendirilmiş bir Excel (.xlsx) dosyasına yazar.

Çıktı, kategoriye göre gruplanmış şekilde düzenlenir: önce "Formlar" grubu,
ardından "Sözleşmeler" grubu, varsa en sonda "Diğer" grubu. Her grup kendi
başlık satırıyla başlar; grup içinde sadece Başlık ve Bağlantı sütunları vardır.
"""

from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scraper import Belge

# Gruplar bu sırayla yazılır.
GRUP_SIRASI = ["Form", "Sözleşme", "Diğer"]
GRUP_BASLIKLARI = {"Form": "Formlar", "Sözleşme": "Sözleşmeler", "Diğer": "Diğer"}

SUTUN_BASLIKLARI = ["Başlık", "Bağlantı (URL)"]

_UST_BILGI_YAZI = Font(italic=True, color="666666", size=9)
_GRUP_DOLGU = PatternFill("solid", fgColor="1F4E78")
_GRUP_YAZI = Font(color="FFFFFF", bold=True, size=12)
_SUTUN_DOLGU = PatternFill("solid", fgColor="D9E2F3")
_SUTUN_YAZI = Font(color="1F4E78", bold=True, size=10)
_KENAR = Border(*[Side(style="thin", color="D0D0D0")] * 4)


def excele_yaz(belgeler: list[Belge], cikti_yolu: str, kaynak_url: str = "") -> str:
    """Belgeleri, kategoriye göre gruplanmış şekilde Excel'e yazar ve dosya yolunu döndürür."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sözleşmeler ve Formlar"

    sutun_sayisi = len(SUTUN_BASLIKLARI)

    # Üst bilgi satırı (kaynak + tarih)
    ust_bilgi = f"Kaynak: {kaynak_url}    |    Oluşturulma: {datetime.now():%d.%m.%Y %H:%M}"
    ws.append([ust_bilgi])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sutun_sayisi)
    ws.cell(row=1, column=1).font = _UST_BILGI_YAZI
    ws.append([])  # boş satır

    belgeler_kategoriye_gore: dict[str, list[Belge]] = {grup: [] for grup in GRUP_SIRASI}
    for belge in belgeler:
        belgeler_kategoriye_gore.setdefault(belge.kategori, []).append(belge)

    for grup in GRUP_SIRASI:
        grup_belgeleri = belgeler_kategoriye_gore.get(grup, [])
        if not grup_belgeleri:
            continue

        # Grup başlık satırı (ör. "Formlar")
        grup_satiri = ws.max_row + 1
        ws.cell(row=grup_satiri, column=1, value=GRUP_BASLIKLARI[grup])
        ws.merge_cells(start_row=grup_satiri, start_column=1, end_row=grup_satiri, end_column=sutun_sayisi)
        g = ws.cell(row=grup_satiri, column=1)
        g.fill = _GRUP_DOLGU
        g.font = _GRUP_YAZI
        g.alignment = Alignment(horizontal="left", vertical="center")

        # Sütun başlıkları (Başlık / Bağlantı)
        sutun_satiri = ws.max_row + 1
        ws.append(SUTUN_BASLIKLARI)
        for sutun in range(1, sutun_sayisi + 1):
            h = ws.cell(row=sutun_satiri, column=sutun)
            h.fill = _SUTUN_DOLGU
            h.font = _SUTUN_YAZI
            h.border = _KENAR

        # Belge satırları
        for belge in grup_belgeleri:
            ws.append([belge.baslik, belge.url])
            r = ws.max_row
            link_hucre = ws.cell(row=r, column=2)
            link_hucre.hyperlink = belge.url
            link_hucre.font = Font(color="0563C1", underline="single")
            for sutun in range(1, sutun_sayisi + 1):
                c = ws.cell(row=r, column=sutun)
                c.border = _KENAR
                c.alignment = Alignment(vertical="center")

        ws.append([])  # gruplar arası boşluk

    # Sütun genişlikleri
    genislikler = {1: 65, 2: 75}
    for sutun, gen in genislikler.items():
        ws.column_dimensions[get_column_letter(sutun)].width = gen

    wb.save(cikti_yolu)
    return cikti_yolu
