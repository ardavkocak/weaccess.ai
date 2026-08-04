"""Envanter uygulamasi view'lari: CRUD, arama/filtreleme, zimmetleme ve belgeler."""
from functools import partial

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from django.db.models import ProtectedError, Q
from django.http import (
    HttpResponse,
    HttpResponseForbidden,
    HttpResponseNotFound,
    HttpResponseServerError,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.generic import CreateView, DeleteView, DetailView, ListView, UpdateView, View

from common.utils import get_client_ip

from . import pdf, services
from .forms import (
    AssignmentCreateForm,
    AssignmentItemFormSet,
    AssignmentReturnForm,
    AssignmentSearchForm,
    CompanyForm,
    DeviceForm,
    DeviceImportForm,
    DeviceSearchForm,
    EmployeeForm,
    EmployeeSearchForm,
)
from .mixins import AdminRequiredMixin
from .models import ActivityLog, Assignment, Company, Device, Employee, Notification

PAGE_SIZE = 15


# ---------------------------------------------------------------------------
# Hata sayfalari
# ---------------------------------------------------------------------------
def error_403_view(request, exception=None):
    return HttpResponseForbidden(render(request, "errors/403.html").content)


def error_404_view(request, exception=None):
    return HttpResponseNotFound(render(request, "errors/404.html").content)


def error_500_view(request):
    return HttpResponseServerError(render(request, "errors/500.html").content)


# ---------------------------------------------------------------------------
# Sirketler
# ---------------------------------------------------------------------------
class CompanyListView(AdminRequiredMixin, ListView):
    model = Company
    template_name = "inventory/company_list.html"
    context_object_name = "companies"
    paginate_by = PAGE_SIZE

    def get_queryset(self):
        qs = Company.objects.all()
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(name__icontains=q)
        return qs.order_by("name")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_query"] = self.request.GET.get("q", "")
        return context


class CompanyCreateView(AdminRequiredMixin, CreateView):
    model = Company
    form_class = CompanyForm
    template_name = "inventory/generic_form.html"
    success_url = reverse_lazy("inventory:company-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Yeni Sirket Ekle"
        context["cancel_url"] = self.success_url
        return context

    def form_valid(self, form):
        messages.success(self.request, "Sirket basariyla olusturuldu.")
        return super().form_valid(form)


class CompanyUpdateView(AdminRequiredMixin, UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = "inventory/generic_form.html"
    success_url = reverse_lazy("inventory:company-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Sirket Duzenle: {self.object.name}"
        context["cancel_url"] = self.success_url
        return context

    def form_valid(self, form):
        messages.success(self.request, "Sirket basariyla guncellendi.")
        return super().form_valid(form)


class CompanyDeleteView(AdminRequiredMixin, DeleteView):
    model = Company
    template_name = "inventory/confirm_delete.html"
    success_url = reverse_lazy("inventory:company-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Sirketi Sil"
        context["cancel_url"] = self.success_url
        return context

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, "Bu sirkete bagli calisanlar oldugu icin silinemez.")
            return redirect(self.success_url)
        messages.success(self.request, "Sirket basariyla silindi.")
        return response


# ---------------------------------------------------------------------------
# Calisanlar
# ---------------------------------------------------------------------------
class EmployeeListView(AdminRequiredMixin, ListView):
    model = Employee
    template_name = "inventory/employee_list.html"
    context_object_name = "employees"
    paginate_by = PAGE_SIZE

    def get_queryset(self):
        qs = Employee.objects.select_related("company").order_by("first_name", "last_name")
        self.search_form = EmployeeSearchForm(self.request.GET)
        if self.search_form.is_valid():
            q = self.search_form.cleaned_data.get("q")
            company = self.search_form.cleaned_data.get("company")
            if q:
                qs = qs.filter(
                    Q(first_name__icontains=q)
                    | Q(last_name__icontains=q)
                    | Q(email__icontains=q)
                    | Q(tc_kimlik_no__icontains=q)
                )
            if company:
                qs = qs.filter(company=company)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_form"] = self.search_form
        return context


class EmployeeDetailView(AdminRequiredMixin, DetailView):
    model = Employee
    template_name = "inventory/employee_detail.html"
    context_object_name = "employee"

    def get_queryset(self):
        return Employee.objects.select_related("company")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee = self.object
        context["active_assignments"] = employee.assignments.filter(returned=False).select_related("device")
        context["assignment_history"] = employee.assignments.filter(returned=True).select_related("device")[:20]
        return context


class EmployeeCreateView(AdminRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = "inventory/generic_form.html"
    success_url = reverse_lazy("inventory:employee-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Yeni Calisan Ekle"
        context["cancel_url"] = self.success_url
        context["is_multipart"] = True
        return context

    def form_valid(self, form):
        """
        Calisani kaydeder. Sistem tamamen admin odakli oldugu icin (personel
        girisi kaldirildi) calisanlar icin artik otomatik giris hesabi
        olusturulmaz — Employee kaydi yalnizca zimmet/Yemek Sistemi/IK
        otomasyonu gibi is mantigi icin kullanilir.
        """
        response = super().form_valid(form)
        self.object.refresh_from_db()

        services.notify_admins(
            "Yeni Personel Eklendi",
            f"{self.object.full_name} sisteme yeni calisan olarak eklendi.",
            link=self.object.get_absolute_url(),
        )
        messages.success(self.request, "Calisan basariyla olusturuldu.")
        return response


class EmployeeUpdateView(AdminRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = "inventory/generic_form.html"
    success_url = reverse_lazy("inventory:employee-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Calisan Duzenle: {self.object.full_name}"
        context["cancel_url"] = self.success_url
        context["is_multipart"] = True
        return context

    def form_valid(self, form):
        messages.success(self.request, "Calisan bilgileri basariyla guncellendi.")
        return super().form_valid(form)


class EmployeeDeleteView(AdminRequiredMixin, DeleteView):
    model = Employee
    template_name = "inventory/confirm_delete.html"
    success_url = reverse_lazy("inventory:employee-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Calisani Sil"
        context["cancel_url"] = self.success_url
        return context

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, "Bu calisana ait zimmet kayitlari oldugu icin silinemez.")
            return redirect(self.success_url)
        messages.success(self.request, "Calisan basariyla silindi.")
        return response


# ---------------------------------------------------------------------------
# Cihazlar (stok)
# ---------------------------------------------------------------------------
class DeviceListView(AdminRequiredMixin, ListView):
    model = Device
    template_name = "inventory/device_list.html"
    context_object_name = "devices"
    paginate_by = PAGE_SIZE

    def get_queryset(self):
        qs = Device.objects.with_stock_counts().order_by("name")
        self.search_form = DeviceSearchForm(self.request.GET)
        if self.search_form.is_valid() and self.search_form.cleaned_data.get("q"):
            qs = qs.filter(name__icontains=self.search_form.cleaned_data["q"])
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_form"] = self.search_form
        return context


class DeviceDetailView(AdminRequiredMixin, DetailView):
    model = Device
    template_name = "inventory/device_detail.html"
    context_object_name = "device"

    def get_queryset(self):
        return Device.objects.with_stock_counts()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        device = self.object
        context["active_assignments"] = device.assignments.filter(returned=False).select_related(
            "employee__company"
        )
        context["assignment_history"] = device.assignments.filter(returned=True).select_related(
            "employee__company"
        )[:20]
        return context


class DeviceCreateView(AdminRequiredMixin, CreateView):
    model = Device
    form_class = DeviceForm
    template_name = "inventory/generic_form.html"
    success_url = reverse_lazy("inventory:device-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Yeni Cihaz Ekle"
        context["cancel_url"] = self.success_url
        return context

    def form_valid(self, form):
        response = super().form_valid(form)
        services.notify_admins(
            "Yeni Cihaz Eklendi",
            f"{self.object} stoga eklendi ({self.object.total_quantity} adet).",
            link=self.object.get_absolute_url(),
        )
        messages.success(self.request, "Cihaz basariyla olusturuldu.")
        return response


class DeviceUpdateView(AdminRequiredMixin, UpdateView):
    model = Device
    form_class = DeviceForm
    template_name = "inventory/generic_form.html"
    success_url = reverse_lazy("inventory:device-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = f"Cihaz Duzenle: {self.object}"
        context["cancel_url"] = self.success_url
        return context

    def form_valid(self, form):
        messages.success(self.request, "Cihaz basariyla guncellendi.")
        return super().form_valid(form)


class DeviceDeleteView(AdminRequiredMixin, DeleteView):
    model = Device
    template_name = "inventory/confirm_delete.html"
    success_url = reverse_lazy("inventory:device-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Cihazi Sil"
        context["cancel_url"] = self.success_url
        return context

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
        except ProtectedError:
            messages.error(self.request, "Bu cihaza ait zimmet kayitlari oldugu icin silinemez.")
            return redirect(self.success_url)
        messages.success(self.request, "Cihaz basariyla silindi.")
        return response


# ---------------------------------------------------------------------------
# Zimmetler (Assignment)
# ---------------------------------------------------------------------------
class AssignmentListView(AdminRequiredMixin, ListView):
    model = Assignment
    template_name = "inventory/assignment_list.html"
    context_object_name = "assignments"
    paginate_by = PAGE_SIZE

    def get_queryset(self):
        qs = Assignment.objects.select_related("employee", "device").order_by("-assigned_date", "-created_at")
        self.search_form = AssignmentSearchForm(self.request.GET)
        if self.search_form.is_valid():
            data = self.search_form.cleaned_data
            if data.get("q"):
                qs = qs.filter(
                    Q(employee__first_name__icontains=data["q"])
                    | Q(employee__last_name__icontains=data["q"])
                    | Q(device__name__icontains=data["q"])
                )
            status = data.get("status")
            if status == "active":
                qs = qs.filter(returned=False)
            elif status == "returned":
                qs = qs.filter(returned=True)
            elif status == "overdue":
                qs = qs.filter(returned=False, expected_return_date__lt=timezone.localdate())
            if data.get("date_from"):
                qs = qs.filter(assigned_date__gte=data["date_from"])
            if data.get("date_to"):
                qs = qs.filter(assigned_date__lte=data["date_to"])
            if data.get("return_date_from"):
                qs = qs.filter(returned_date__gte=data["return_date_from"])
            if data.get("return_date_to"):
                qs = qs.filter(returned_date__lte=data["return_date_to"])
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["search_form"] = self.search_form
        return context


class AssignmentDetailView(AdminRequiredMixin, DetailView):
    model = Assignment
    template_name = "inventory/assignment_detail.html"
    context_object_name = "assignment"

    def get_queryset(self):
        return Assignment.objects.select_related("employee__company", "device")


class AssignmentCreateView(AdminRequiredMixin, View):
    """Yeni zimmet olusturma: ayni islemde birden fazla cihaz zimmetlenebilir.

    Is mantigi services.assign_devices icinde toplanmistir; stok kontrolu ve
    kayit olusturma tek bir transaction icinde yapilir.
    """

    template_name = "inventory/assignment_form.html"
    title = "Yeni Zimmet Olustur"

    def _render(self, request, form, formset):
        return render(
            request,
            self.template_name,
            {"form": form, "formset": formset, "title": self.title},
        )

    def get(self, request):
        # Cihaz detayindan gelindiyse ilk satir o cihazla doldurulur.
        device_id = request.GET.get("device")
        formset_initial = [{"device": device_id}] if device_id else None
        return self._render(
            request,
            AssignmentCreateForm(initial={"employee": request.GET.get("employee")}),
            AssignmentItemFormSet(initial=formset_initial),
        )

    def post(self, request):
        form = AssignmentCreateForm(request.POST)
        formset = AssignmentItemFormSet(request.POST)
        if not (form.is_valid() and formset.is_valid()):
            return self._render(request, form, formset)

        items = [
            (item_form.cleaned_data["device"], item_form.cleaned_data["quantity"])
            for item_form in formset
            if item_form.cleaned_data.get("device")
        ]
        if not items:
            messages.error(request, "En az bir cihaz secmelisiniz.")
            return self._render(request, form, formset)

        try:
            assignments = services.assign_devices(
                employee=form.cleaned_data["employee"],
                items=items,
                assigned_by=request.user,
                assigned_date=form.cleaned_data["assigned_date"],
                expected_return_date=form.cleaned_data["expected_return_date"],
                notes=form.cleaned_data["notes"],
                ip_address=get_client_ip(request),
            )
        except ValidationError as exc:
            messages.error(request, "; ".join(exc.messages))
            return self._render(request, form, formset)

        # Yeni zimmette yalnizca teslim tutanagi uretilir (iade tutanagi degil).
        total = sum(quantity for _, quantity in items)
        messages.success(
            request,
            f"{len(items)} kalem / {total} adet cihaz basariyla zimmetlendi. Teslim tutanagi olusturuldu.",
        )
        return redirect("inventory:assignment-detail", pk=assignments[0].pk)


class AssignmentReturnView(AdminRequiredMixin, View):
    """Bir zimmet kaydinin iade islemini onaylayan view."""

    template_name = "inventory/assignment_return_confirm.html"

    def get(self, request, pk):
        assignment = get_object_or_404(Assignment, pk=pk, returned=False)
        form = AssignmentReturnForm(instance=assignment)
        return render(request, self.template_name, {"assignment": assignment, "form": form})

    def post(self, request, pk):
        assignment = get_object_or_404(Assignment, pk=pk, returned=False)
        form = AssignmentReturnForm(request.POST, instance=assignment)
        if form.is_valid():
            try:
                services.return_device(
                    assignment=assignment,
                    returned_by=request.user,
                    returned_date=form.cleaned_data.get("returned_date"),
                    return_notes=form.cleaned_data["return_notes"],
                    return_condition=form.cleaned_data["return_condition"],
                    damage_description=form.cleaned_data["damage_description"],
                    ip_address=get_client_ip(request),
                )
            except ValidationError as exc:
                messages.error(request, "; ".join(exc.messages))
            else:
                # Iade isleminde teslim tutanagi degil, iade tutanagi uretilir.
                messages.success(request, "Cihaz iadesi basariyla alindi. Iade tutanagi olusturuldu.")
                return redirect("inventory:assignment-detail", pk=assignment.pk)
        return render(request, self.template_name, {"assignment": assignment, "form": form})


class AssignmentDeleteView(AdminRequiredMixin, DeleteView):
    """Yalnizca iade edilmis (kapanmis) zimmet kayitlarinin silinmesine izin verir."""

    model = Assignment
    template_name = "inventory/confirm_delete.html"
    success_url = reverse_lazy("inventory:assignment-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Zimmet Kaydini Sil"
        context["cancel_url"] = self.success_url
        return context

    def form_valid(self, form):
        if not self.object.returned:
            messages.error(self.request, "Aktif bir zimmet kaydi silinemez. Once iade alinmalidir.")
            return redirect(self.success_url)
        messages.success(self.request, "Zimmet kaydi basariyla silindi.")
        return super().form_valid(form)


# ---------------------------------------------------------------------------
# Bildirimler
# ---------------------------------------------------------------------------
@login_required
def notification_list_view(request):
    """Kullanicinin tum bildirimlerini listeledigi sayfa."""
    notifications = Notification.objects.filter(user=request.user).order_by("-created_at")
    return render(request, "inventory/notification_list.html", {"notifications": notifications})


@login_required
@require_POST
def mark_notification_read(request, pk):
    """Tek bir bildirimi okundu olarak isaretler (AJAX)."""
    notification = get_object_or_404(Notification, pk=pk, user=request.user)
    notification.is_read = True
    notification.save(update_fields=["is_read"])
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    return redirect(notification.link or "inventory:notification-list")


@login_required
@require_POST
def mark_all_notifications_read(request):
    """Kullanicinin tum okunmamis bildirimlerini okundu olarak isaretler."""
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"success": True})
    messages.success(request, "Tum bildirimler okundu olarak isaretlendi.")
    return redirect("inventory:notification-list")


# ---------------------------------------------------------------------------
# Global Canli Arama (Navbar)
# ---------------------------------------------------------------------------
@login_required
def global_search_api(request):
    """Cihaz ve calisanlarda anlik arama yapan JSON API.

    Navbar'daki arama kutusundan JavaScript ile cagrilir (bkz. static/js/main.js).
    """
    query = request.GET.get("q", "").strip()
    results = []

    if len(query) >= 2:
        devices = Device.objects.with_stock_counts().filter(name__icontains=query)[:5]
        for device in devices:
            results.append(
                {
                    "type": "Cihaz",
                    "icon": "bi-box-seam",
                    "title": device.name,
                    "subtitle": f"Bosta {device.available_count} / Toplam {device.total_quantity}",
                    "url": device.get_absolute_url(),
                }
            )

        employees = Employee.objects.select_related("company").filter(
            Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(email__icontains=query)
        )[:5]
        for employee in employees:
            results.append(
                {
                    "type": "Calisan",
                    "icon": "bi-person",
                    "title": employee.full_name,
                    "subtitle": employee.company.name,
                    "url": employee.get_absolute_url(),
                }
            )

    return JsonResponse({"results": results})


def _admin_required(request):
    """Fonksiyon tabanli view'larda admin yetkisini dogrular; degilse yonlendirme dondurur."""
    if not request.user.is_authenticated:
        return redirect(f"{reverse_lazy('accounts:login')}?next={request.path}")
    if not request.user.is_admin_role:
        messages.error(request, "Bu islem icin yetkiniz bulunmuyor.")
        return redirect("dashboard:home")
    return None


# ---------------------------------------------------------------------------
# Excel Disa / Ice Aktarma
# ---------------------------------------------------------------------------
def device_export_excel(request):
    """Mevcut filtrelerle eslesen cihaz stogunu .xlsx dosyasi olarak disa aktarir."""
    guard = _admin_required(request)
    if guard:
        return guard

    from openpyxl import Workbook
    from openpyxl.styles import Font

    list_view = DeviceListView()
    list_view.request = request
    devices = list_view.get_queryset()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cihazlar"

    sheet.append(["Cihaz Adi", "Toplam Adet", "Zimmette", "Bosta"])
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for device in devices:
        sheet.append([device.name, device.total_quantity, device.assigned_count, device.available_count])

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 4, 40)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="cihazlar.xlsx"'
    workbook.save(response)

    services.log_activity(
        request.user,
        ActivityLog.ActionType.OTHER,
        "Cihaz listesi Excel olarak disa aktarildi.",
        get_client_ip(request),
    )
    return response


def device_import_excel(request):
    """Yuklenen .xlsx dosyasindan toplu cihaz stogu olusturur."""
    guard = _admin_required(request)
    if guard:
        return guard

    if request.method == "POST":
        form = DeviceImportForm(request.POST, request.FILES)
        if form.is_valid():
            created_count, skipped = _process_device_import(form.cleaned_data["excel_file"])
            if created_count:
                messages.success(request, f"{created_count} cihaz basariyla eklendi.")
                services.log_activity(
                    request.user,
                    ActivityLog.ActionType.CREATE,
                    f"Excel'den {created_count} cihaz ice aktarildi.",
                    get_client_ip(request),
                )
            if skipped:
                messages.warning(request, f"{len(skipped)} satir atlandi: " + "; ".join(skipped[:5]))
            return redirect("inventory:device-list")
    else:
        form = DeviceImportForm()

    return render(request, "inventory/device_import.html", {"form": form})


def _process_device_import(excel_file):
    """Excel dosyasindaki satirlari okuyup Device stok kayitlarina donusturur.

    Returns:
        tuple: (basariyla_olusturulan_sayisi, atlanan_satirlarin_aciklamalari)
    """
    from openpyxl import load_workbook

    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    sheet = workbook.active

    created_count = 0
    skipped = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell in (None, "") for cell in row):
            continue

        name = str(row[0]).strip() if row[0] else ""
        if not name:
            skipped.append(f"Satir {index}: cihaz adi bos")
            continue
        if Device.objects.filter(name__iexact=name).exists():
            skipped.append(f"Satir {index}: '{name}' zaten kayitli")
            continue

        try:
            quantity = int(row[1]) if len(row) > 1 and row[1] not in (None, "") else 1
        except (TypeError, ValueError):
            skipped.append(f"Satir {index}: toplam adet sayi olmali")
            continue
        if quantity < 0:
            skipped.append(f"Satir {index}: toplam adet negatif olamaz")
            continue

        Device.objects.create(name=name, total_quantity=quantity)
        created_count += 1

    return created_count, skipped


# ---------------------------------------------------------------------------
# PDF Tutanaklar (teslim / iade)
# ---------------------------------------------------------------------------
def _render_assignment_pdf(request, assignment, builder, file_prefix, log_message):
    """Verilen tutanak olusturucusunu calistirip PDF yanitini dondurur."""
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="{file_prefix}_{assignment.pk:06d}.pdf"'
    builder(response, assignment)

    services.log_activity(
        request.user,
        ActivityLog.ActionType.OTHER,
        f"{assignment} icin {log_message} olusturuldu.",
        get_client_ip(request),
    )
    return response


def _get_assignment_for_pdf(pk):
    return get_object_or_404(
        Assignment.objects.select_related("employee__company", "device", "assigned_by", "returned_by"),
        pk=pk,
    )


def assignment_delivery_pdf_view(request, pk):
    """Cihaz teslim edilirken olusturulan zimmet teslim tutanagi (PDF)."""
    guard = _admin_required(request)
    if guard:
        return guard

    assignment = _get_assignment_for_pdf(pk)
    return _render_assignment_pdf(
        request, assignment, pdf.build_delivery_pdf, "zimmet_teslim_tutanagi", "PDF zimmet teslim tutanagi"
    )


def assignment_return_pdf_view(request, pk):
    """Cihaz iade alindiginda olusturulan iade teslim tutanagi (PDF).

    Teslim tutanagindan tamamen farkli bir tasarima sahiptir ve yalnizca iade
    edilmis zimmet kayitlari icin uretilebilir.
    """
    guard = _admin_required(request)
    if guard:
        return guard

    assignment = _get_assignment_for_pdf(pk)
    if not assignment.returned:
        messages.error(request, "Iade tutanagi yalnizca iade alinmis zimmet kayitlari icin olusturulabilir.")
        return redirect(assignment.get_absolute_url())

    # Insan Kaynaklari imza bloguna tutanagi olusturan yetkilinin adi yazilir.
    return _render_assignment_pdf(
        request,
        assignment,
        partial(pdf.build_return_pdf, hr_user=request.user),
        "iade_teslim_tutanagi",
        "PDF iade teslim tutanagi",
    )
