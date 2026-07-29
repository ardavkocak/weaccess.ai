"""
Excel (.xlsx) yemek listesi okuyucu — mealImport.service.js'in Python portu.

Iki format desteklenir (bkz. orijinal JS docstring'i, ayni mantik):
  1) Haftalik blok: gun adlari sutun basligi, altinda tarih + yemek satirlari.
  2) Duz tablo: her satir bir gun, "Tarih" + yemek sutunlari.

openpyxl kullanir (JS tarafinda ExcelJS'in karsiligi).
"""
from __future__ import annotations

import re
from datetime import date, timedelta

from openpyxl import load_workbook

from .date_utils import today_iso
from .text_utils import normalize_tr

HEADER_SEARCH_ROWS = 15
MAX_ROWS = 1000

HEADER_ALIASES = {
    "date": ["tarih", "date", "gun tarihi"],
    "day": ["gun", "day", "haftanin gunu", "gunu"],
}

MEAL_HINTS = [
    "menu", "yemek", "corba", "ana", "yan", "salata", "tatli", "icecek",
    "ogun", "ogle", "aksam", "kahvalti", "icerik", "liste",
]

WEEKDAYS = {
    "pazartesi": "Pazartesi", "sali": "Salı", "carsamba": "Çarşamba",
    "persembe": "Perşembe", "cuma": "Cuma", "cumartesi": "Cumartesi", "pazar": "Pazar",
}

MIN_WEEKDAYS_IN_HEADER = 2


def _weekday_of(text):
    return WEEKDAYS.get(normalize_tr(text))


def _cell_text(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if hasattr(value, "isoformat"):  # date/datetime
        return value.isoformat()
    return str(value).strip()


def _parse_date(cell):
    value = cell.value
    if hasattr(value, "year") and hasattr(value, "month"):  # date/datetime
        return date(value.year, value.month, value.day).isoformat()

    text = _cell_text(cell).strip()
    if not text:
        return None

    iso = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if iso:
        return f"{iso.group(1)}-{iso.group(2)}-{iso.group(3)}"

    dmy = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$", text)
    if dmy:
        day, month, year = int(dmy.group(1)), int(dmy.group(2)), int(dmy.group(3))
        if year < 100:
            year += 2000
        if not (1 <= month <= 12 and 1 <= day <= 31):
            return None
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return None

    try:
        serial = float(text)
    except ValueError:
        return None
    if 1 < serial < 100_000:
        base = date(1899, 12, 30)  # Excel 1900 epoch (1904 sistemi haric)
        return (base + timedelta(days=serial)).isoformat()
    return None


def _matches_alias(header, aliases):
    normalized = normalize_tr(header)
    if not normalized:
        return False
    return any(normalized == alias or alias in normalized for alias in aliases)


def _split_dishes(text):
    parts = re.split(r"[\n\r;]+|,(?=\s)", str(text or ""))
    return [re.sub(r"\s+", " ", p.strip()) for p in parts if p.strip()]


def _parse_kcal(cell):
    text = _cell_text(cell).strip()
    if not text:
        return None
    digits = re.sub(r"[.\s]", "", text)
    match = re.search(r"\d+", digits)
    if not match:
        return None
    value = int(match.group(0))
    return value if value > 0 else None


def _to_dish_entries(dish_text, kcal):
    dishes = _split_dishes(dish_text)
    if len(dishes) == 1:
        return [{"name": dishes[0], "kcal": kcal}]
    return [{"name": name, "kcal": None} for name in dishes]


def _is_dish_text(text):
    trimmed = str(text or "").strip()
    if not trimmed:
        return False
    if re.match(r"^[\d.,\s]+$", trimmed):
        return False
    if _weekday_of(trimmed):
        return False
    if normalize_tr(trimmed) == "enerji":
        return False
    return True


def _detect_columns(ws):
    last_row = min(ws.max_row, HEADER_SEARCH_ROWS)
    for row_number in range(1, last_row + 1):
        headers = []
        for cell in ws[row_number]:
            text = _cell_text(cell).strip()
            if text:
                headers.append((cell.column, text))
        if len(headers) < 2:
            continue

        date_col = next((c for c, t in headers if _matches_alias(t, HEADER_ALIASES["date"])), None)
        if date_col is None:
            continue

        day_col = next((c for c, t in headers if c != date_col and _matches_alias(t, HEADER_ALIASES["day"])), None)
        item_cols = [c for c, _ in headers if c != date_col and c != day_col]
        if not item_cols:
            continue

        labels = {c: t for c, t in headers}
        return {
            "header_row": row_number,
            "columns": {
                "date": date_col, "day": day_col, "items": item_cols, "labels": labels,
            },
        }
    return None


def _read_day_header_row(ws, row_number):
    day_columns = []
    for cell in ws[row_number]:
        day_name = _weekday_of(_cell_text(cell))
        if day_name:
            day_columns.append({"col": cell.column, "day_name": day_name})
    return day_columns if len(day_columns) >= MIN_WEEKDAYS_IN_HEADER else None


def _parse_block_layout(ws):
    headers = []
    for row_number in range(1, ws.max_row + 1):
        day_columns = _read_day_header_row(ws, row_number)
        if day_columns:
            headers.append({"row_number": row_number, "day_columns": day_columns})

    if not headers:
        return None

    rows = []
    seen = set()
    warnings = []
    skipped = 0

    for index, header in enumerate(headers):
        block_end = headers[index + 1]["row_number"] - 1 if index + 1 < len(headers) else ws.max_row
        day_cols = {d["col"] for d in header["day_columns"]}

        energy_cols = []
        for cell in ws[header["row_number"]]:
            text = _cell_text(cell).strip()
            if not text or cell.column in day_cols:
                continue
            if normalize_tr(text) == "enerji":
                energy_cols.append(cell.column)

        def energy_col_for(day_col_number):
            candidates = sorted(c for c in energy_cols if c > day_col_number)
            return candidates[0] if candidates else None

        date_row = None
        for row_number in range(header["row_number"] + 1, block_end + 1):
            if any(_parse_date(ws.cell(row=row_number, column=d["col"])) for d in header["day_columns"]):
                date_row = row_number
                break

        if date_row is None:
            warnings.append(f'{header["row_number"]}. satırdaki gün başlığının altında tarih bulunamadı; blok atlandı.')
            continue

        dish_rows = []
        for row_number in range(date_row + 1, block_end + 1):
            has_any = any(_cell_text(ws.cell(row=row_number, column=d["col"])).strip() for d in header["day_columns"])
            if not has_any:
                break
            dish_rows.append(row_number)

        for day in header["day_columns"]:
            menu_date = _parse_date(ws.cell(row=date_row, column=day["col"]))
            if not menu_date:
                continue

            energy_col = energy_col_for(day["col"])
            items = []
            for row_number in dish_rows:
                text = _cell_text(ws.cell(row=row_number, column=day["col"]))
                if not _is_dish_text(text):
                    continue
                kcal = _parse_kcal(ws.cell(row=row_number, column=energy_col)) if energy_col else None
                items.extend(_to_dish_entries(text, kcal))

            if not items:
                skipped += 1
                continue
            if menu_date in seen:
                warnings.append(f"{menu_date} tarihi dosyada birden fazla kez var; ilk blok kullanıldı.")
                continue
            seen.add(menu_date)
            rows.append({"menu_date": menu_date, "day_name": day["day_name"], "items": items})

    if not rows:
        return None
    return {"rows": rows, "skipped": skipped, "warnings": warnings, "columns": {"format": "block", "block_count": len(headers)}}


def _parse_table_layout(ws, columns, header_row):
    rows = []
    seen = set()
    warnings = []
    skipped = 0

    for row_number in range(header_row + 1, ws.max_row + 1):
        if len(rows) >= MAX_ROWS:
            warnings.append(f"Dosyada {MAX_ROWS} satırdan fazlası var; fazlası okunmadı.")
            break

        menu_date = _parse_date(ws.cell(row=row_number, column=columns["date"]))
        if not menu_date:
            has_content = any(_cell_text(ws.cell(row=row_number, column=c)).strip() for c in columns["items"])
            if has_content:
                skipped += 1
            continue

        items = []
        for col in columns["items"]:
            for name in _split_dishes(_cell_text(ws.cell(row=row_number, column=col))):
                items.append({"name": name, "kcal": None})

        if not items:
            skipped += 1
            continue
        if menu_date in seen:
            warnings.append(f"{menu_date} tarihi dosyada birden fazla kez var; ilk satır kullanıldı.")
            continue
        seen.add(menu_date)

        day_name = _cell_text(ws.cell(row=row_number, column=columns["day"])).strip() if columns.get("day") else ""
        rows.append({"menu_date": menu_date, "day_name": day_name, "items": items})

    return {"rows": rows, "skipped": skipped, "warnings": warnings, "columns": {**columns, "format": "table"}, "header_row": header_row}


def parse_workbook(file_obj, sheet_name=None):
    """
    Yuklenen .xlsx dosyasini okur.

    Donus: {"ok": bool, "message": str|None, "rows": [...], "warnings": [...], "columns": {...}}
    """
    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "message": f"Dosya okunamadı: {exc}", "rows": [], "warnings": []}

    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
    if ws is None:
        return {"ok": False, "message": "Dosyada okunabilir bir sayfa bulunamadı.", "rows": [], "warnings": []}

    block = _parse_block_layout(ws)
    parsed = block
    if parsed is None:
        detected = _detect_columns(ws)
        if detected is None:
            return {
                "ok": False,
                "message": (
                    "Dosya biçimi tanınamadı. İki düzen desteklenir: (1) haftalık blok — gün "
                    "adları sütun başlığı; (2) düz tablo — bir Tarih sütunu ve yemek sütunları."
                ),
                "rows": [], "warnings": [],
            }
        parsed = _parse_table_layout(ws, detected["columns"], detected["header_row"])

    if not parsed["rows"]:
        return {
            "ok": False,
            "message": "Dosyada okunabilir menü satırı bulunamadı. Tarih ve yemek sütunlarını kontrol edin.",
            "rows": [], "warnings": parsed["warnings"], "columns": parsed["columns"],
        }

    today = today_iso()
    past = sum(1 for r in parsed["rows"] if r["menu_date"] < today)
    warnings = list(parsed["warnings"])
    if past:
        warnings.append(f"{past} gün geçmiş tarihli.")

    return {
        "ok": True,
        "rows": sorted(parsed["rows"], key=lambda r: r["menu_date"]),
        "warnings": warnings,
        "columns": parsed["columns"],
        "skipped": parsed["skipped"],
    }
