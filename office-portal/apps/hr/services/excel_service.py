"""
Excel/CSV okuyucu — ik-otomasyon/src/services/excel.service.js'in Python portu.

.xlsx ve .csv tam destekli. Eski .xls (binary) formati bu portta
desteklenmez; kullaniciya acik bir hata mesaji gosterilir (bkz. views.py).
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime

from openpyxl import load_workbook

HEADER_ALIASES = {
    "full_name": ["ad soyad", "adsoyad", "çalışan adı", "personel adı", "isim soyisim", "isim", "ad", "soyad", "adı"],
    "birth_date": ["doğum tarihi", "doğumtarihi", "dogum tarihi", "dogumtarihi", "doğum", "dogum", "birth", "birthdate", "doğum günü"],
    "hire_date": ["işe giriş tarihi", "işegiriş tarihi", "ise giris tarihi", "işe başlama tarihi", "işebaşlama tarihi", "işe giriş", "başlangıç tarihi", "başlangıç", "hire", "start date"],
}


def _normalize(value) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _find_column(headers, aliases):
    normalized = [_normalize(h) for h in headers]
    for alias in aliases:
        key = _normalize(alias)
        if key in normalized:
            return normalized.index(key)
    for alias in aliases:
        key = _normalize(alias)
        for i, h in enumerate(normalized):
            if key in h:
                return i
    return -1


def _parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, (date, datetime)):
        return date(value.year, value.month, value.day)
    text = str(value).strip()
    match = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{4})$", text)
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        try:
            return date(year, month, day)
        except ValueError:
            return None
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def read_employees(uploaded_file):
    """
    Yuklenen dosyayi okur.

    @returns {"headers": [...], "employees": [...], "missing": [...]}
      employees[i] = {"row_number", "values", "full_name", "birth_date", "hire_date"}
      (birth_date/hire_date "YYYY-MM-DD" string ya da None)
    """
    name = (uploaded_file.name or "").lower()

    if name.endswith(".csv"):
        text = uploaded_file.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        matrix = list(reader)
    elif name.endswith(".xlsx"):
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.worksheets[0]
        matrix = [[cell.value for cell in row] for row in ws.iter_rows()]
    else:
        raise ValueError("Bu dosya türü desteklenmiyor. Lütfen .xlsx veya .csv dosyası seçin.")

    if not matrix:
        raise ValueError("Dosya boş görünüyor.")

    headers = [str(h or "").strip() for h in matrix[0]]
    full_name_col = _find_column(headers, HEADER_ALIASES["full_name"])
    birth_col = _find_column(headers, HEADER_ALIASES["birth_date"])
    hire_col = _find_column(headers, HEADER_ALIASES["hire_date"])

    employees = []
    for index, row in enumerate(matrix[1:]):
        if not any(str(v or "").strip() for v in row):
            continue
        full_name = str(row[full_name_col]).strip() if 0 <= full_name_col < len(row) and row[full_name_col] else ""
        if not full_name:
            continue

        birth_date = _parse_date(row[birth_col]) if 0 <= birth_col < len(row) else None
        hire_date = _parse_date(row[hire_col]) if 0 <= hire_col < len(row) else None

        employees.append({
            "row_number": index + 2,
            "values": [row[i] if i < len(row) else "" for i in range(len(headers))],
            "full_name": full_name,
            "birth_date": birth_date.isoformat() if birth_date else None,
            "hire_date": hire_date.isoformat() if hire_date else None,
        })

    missing = []
    if full_name_col < 0:
        missing.append("Ad Soyad")
    if birth_col < 0:
        missing.append("Doğum Tarihi")
    if hire_col < 0:
        missing.append("İşe Giriş Tarihi")

    return {"headers": headers, "employees": employees, "missing": missing}
